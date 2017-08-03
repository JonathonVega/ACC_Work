import openpyxl
import time
import requests
import csv
from bs4 import BeautifulSoup

suppSites = ['http://www.onlinejacc.org/content/meeting-abstract-supplements', 'http://www.interventions.onlinejacc.org/content/meeting-abstract-supplements']

def getScienceDirectLinks():
    scienceDirectLinks = []
    for site in suppSites:
        r = requests.get(site)
        archiveSoup = BeautifulSoup(r.text, 'lxml')
        print(archiveSoup)
        for links in archiveSoup.find_all('a', href=True):
            print(links)
            if('www.sciencedirect.com' in links.get('href')):
                scienceDirectLinks.append(links.get('href'))
    return scienceDirectLinks

#def getIssueSupplementName(link):
    #VI = link[link.index('/journal/') + 18:]
    #issue = VI[VI.index('/') + 1:].replace('/supp/', '_')
    #if(issue.index('_'))

# example: http://www.sciencedirect.com/science/journal/07351097/69/11/supp/S
SDLinksList = getScienceDirectLinks()
print(SDLinksList)
print("cool")
wbOG = openpyxl.load_workbook('SCArticlesAbstractsOnly.xlsx')

print(wbOG.get_sheet_names())
sheetOG = wbOG.get_sheet_by_name('SCArticles')

numOfRows = sheetOG.max_row
numOfColumns = sheetOG.max_column

filepath = 'C:\\Users\\jvega\\Documents\\PythonScripts\\FilteredSCArticlesAbstractsOnly_S.xlsx'
wbNew = openpyxl.Workbook()
wbNew.save(filepath)

wbNew = openpyxl.load_workbook('FilteredSCArticlesAbstractsOnly_S.xlsx')
sheetNew = wbNew.get_sheet_by_name('Sheet')
sheetNew.title = 'SCArticles'

wbNew.create_sheet('SCArticlesCount')
sheetNewCount = wbNew.get_sheet_by_name('SCArticlesCount')
sheetNewCount.cell(row=1, column=1).value = 'Row Labels'
sheetNewCount.cell(row=1, column=2).value = 'Count of key'

# Set Titles in SCArticles sheet
for column in range(1, numOfColumns + 1):
    sheetNew.cell(row=1, column=column).value = sheetOG.cell(row=1, column=column).value

wbNew.save('FilteredSCArticlesAbstractsOnly_S.xlsx')

countRow = 2
for link in SDLinksList:
    VI = link[link.index('/journal/') + 18:link.index('/supp')]
    volume = VI[:VI.index('/')]
    issue = VI[VI.index('/') + 1:]
    print('Volume is:' + volume + '.')
    print('Issue is:' + issue + '.')
    articlesInIssue = 0
    for originalRow in range(2,sheetOG.max_row):
        if(str(sheetOG.cell(row=originalRow, column=5).value) == volume and (issue in str(sheetOG.cell(row=originalRow, column=6).value) and ('s' in str(sheetOG.cell(row=originalRow, column=6).value) or 'S' in str(sheetOG.cell(row=originalRow, column=6).value) ))):
            issue = str(sheetOG.cell(row=originalRow, column=6).value)
            print('Volume is:' + volume + '.')
            print('Issue is:' + issue + '.')
            nextNewRow = sheetNew.max_row + 1
            articlesInIssue += 1
            print('FOUND ONE!!!')
            for column in range(1,8):
                if(column == 7):
                    key = '=CONCATENATE({},"-",{})'.format(str(sheetOG.cell(row=originalRow, column=5).value),str(sheetOG.cell(row=originalRow, column=6).value))
                    sheetNew.cell(row=nextNewRow, column=column).value = key
                    sheetNewCount.cell(row=countRow, column=1).value = key
                    sheetNewCount.cell(row=countRow, column=2).value = articlesInIssue
                    wbNew.save('FilteredSCArticlesAbstractsOnly_S.xlsx')
                else:
                    sheetNew.cell(row=nextNewRow, column=column).value = sheetOG.cell(row=originalRow, column=column).value
                    wbNew.save('FilteredSCArticlesAbstractsOnly_S.xlsx')
    if(articlesInIssue == 0):
        sheetNewCount.cell(row=countRow, column=1).value = '=CONCATENATE({},"-",{})'.format(volume,issue)
        sheetNewCount.cell(row=countRow, column=2).value = articlesInIssue
    countRow += 1

wbNew.save('FilteredSCArticlesAbstractsOnly_S.xlsx')

print('Finished!')

time.sleep(10)
