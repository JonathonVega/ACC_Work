import openpyxl
import time
import requests
import csv
from bs4 import BeautifulSoup
import os

# The two sites to get the supplements from
suppSites = ['http://www.onlinejacc.org/content/meeting-abstract-supplements', 'http://www.interventions.onlinejacc.org/content/meeting-abstract-supplements']

# Returns a list of all the Science Direct links
def getScienceDirectLinks():
    scienceDirectLinks = []
    for site in suppSites:
        r = requests.get(site)
        archiveSoup = BeautifulSoup(r.text, 'lxml')
        print(archiveSoup)
        for links in archiveSoup.find_all('a', href=True):
            if('www.sciencedirect.com' in links.get('href')):
                scienceDirectLinks.append(links.get('href'))
    return scienceDirectLinks

# example: http://www.sciencedirect.com/science/journal/07351097/69/11/supp/S
SDLinksList = getScienceDirectLinks() # Gets the list of Science Direct links
print(SDLinksList)

wbOG = openpyxl.load_workbook('SCArticlesAbstractsOnly.xlsx') # Opens workbook
sheetOG = wbOG.get_sheet_by_name('SCArticles') # Opens worksheet in workbook

numOfRows = sheetOG.max_row
numOfColumns = sheetOG.max_column

filepath = os.getcwd() + '\\' + 'FilteredSCArticlesAbstractsOnly_S.xlsx' # Gets the current filepath of the program and creates a new excel file
wbNew = openpyxl.Workbook() # Opens new workbook
wbNew.save(filepath) # saves the workbook

wbNew = openpyxl.load_workbook('FilteredSCArticlesAbstractsOnly_S.xlsx') # reopens new workbook
sheetNew = wbNew.get_sheet_by_name('Sheet') # opens new worksheet
sheetNew.title = 'SCArticles'

wbNew.create_sheet('SCArticlesCount') # Creates a new worksheet where the count of the abstract supplements will be in
sheetNewCount = wbNew.get_sheet_by_name('SCArticlesCount')
sheetNewCount.cell(row=1, column=1).value = 'Row Labels'
sheetNewCount.cell(row=1, column=2).value = 'Count of key'

# Set Titles in SCArticles sheet
for column in range(1, numOfColumns + 1):
    sheetNew.cell(row=1, column=column).value = sheetOG.cell(row=1, column=column).value

wbNew.save('FilteredSCArticlesAbstractsOnly_S.xlsx')

# Begins retrieving all supplements from the 'SCArticlesAbstractsOnly.xlsx' that are found from the Science Direct links
countRow = 2
for link in SDLinksList:
    VI = link[link.index('/journal/') + 18:link.index('/supp')]
    volume = VI[:VI.index('/')]
    issue = VI[VI.index('/') + 1:]
    print('Now working on Volume ' + volume 'and issue ' + issue)
    print('Issue is:' + issue + '.')
    articlesInIssue = 0
    for originalRow in range(2,sheetOG.max_row):
        # If the issues of the volumes contain an 's' or 'S', then they will be downloaded
        if(str(sheetOG.cell(row=originalRow, column=5).value) == volume and (issue in str(sheetOG.cell(row=originalRow, column=6).value) and ('s' in str(sheetOG.cell(row=originalRow, column=6).value) or 'S' in str(sheetOG.cell(row=originalRow, column=6).value) ))):
            issue = str(sheetOG.cell(row=originalRow, column=6).value)
            print('Volume is:' + volume + '.')
            print('Issue is:' + issue + '.')
            nextNewRow = sheetNew.max_row + 1 # Goes to next empty row
            articlesInIssue += 1 # Keeps count of the amount of supplements in issue to put in 'SCArticlesCount' worksheet
            print('Another found!')
            for column in range(1,8):
                if(column == 7):
                    key = '=CONCATENATE({},"-",{})'.format(str(sheetOG.cell(row=originalRow, column=5).value),str(sheetOG.cell(row=originalRow, column=6).value))
                    sheetNew.cell(row=nextNewRow, column=column).value = key
                    sheetNewCount.cell(row=countRow, column=1).value = key
                    sheetNewCount.cell(row=countRow, column=2).value = articlesInIssue
                    wbNew.save('FilteredSCArticlesAbstractsOnly_S.xlsx')
                else:
                    sheetNew.cell(row=nextNewRow, column=column).value = sheetOG.cell(row=originalRow, column=column).value
                    wbNew.save('FilteredSCArticlesAbstractsOnly_S.xlsx')
    if(articlesInIssue == 0):
        sheetNewCount.cell(row=countRow, column=1).value = '=CONCATENATE({},"-",{})'.format(volume,issue)
        sheetNewCount.cell(row=countRow, column=2).value = articlesInIssue
    countRow += 1

wbNew.save('FilteredSCArticlesAbstractsOnly_S.xlsx')

print('Program done!')

time.sleep(20)
