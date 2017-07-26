import openpyxl
import time

print("Hello, World!")

def getFirstPageFromSheet(sheetMain, sheetJournal):
    continueScript = True
    for row in range(2, 73550):
        articleDOI = sheetMain.cell(row = row, column = 2).value
        print(articleDOI)
        if(continueScript == False):
            break
        for jrow in range(2, 25700):
            jDOI = sheetJournal.cell(row = jrow, column = 1).value
            if(articleDOI == jDOI):
                sheetMain['H' + str(row)] = sheetJournal['D' + str(jrow)].value
                print('FOUND!!!')
                break
            #elif(sheetJournal['D' + str(jrow)].value is None):
                #continueScript = False
            else:
                continue



wbMain = openpyxl.load_workbook('WebsiteToText\SCArticles.xlsx')

print(wbMain.get_sheet_names())
sheetMain = wbMain.get_sheet_by_name('SCArticles')
sheetMain['H1'] = "FirstPageNo"
wbMain.save('SCArticles&JACCJournals1.xlsx')

wbMain = openpyxl.load_workbook('SCArticles&JACCJournals1.xlsx')
sheetMain = wbMain.get_sheet_by_name('SCArticles')

# JACC
wbJournal = openpyxl.load_workbook('WebsiteToText\XJACCJournalData.xlsx')
sheetJournal = wbJournal.get_sheet_by_name('JACCJournalData')
print(sheetJournal['A1'].value)
print('ok')

getFirstPageFromSheet(sheetMain, sheetJournal)
wbMain.save('SCArticles&JACCJournals1.xlsx')
print('Done with JACC')
time.sleep(10)

# INT
wbJournal = openpyxl.load_workbook('WebsiteToText\XINTJournalData.xlsx')
sheetJournal = wbJournal.get_sheet_by_name('INTJournalData')

getFirstPageFromSheet(sheetMain, sheetJournal)
wbMain.save('SCArticles&JACCJournals1.xlsx')
print('Done with INT')
time.sleep(10)

# IMG
wbJournal = openpyxl.load_workbook('WebsiteToText\XIMGJournalData.xlsx')
sheetJournal = wbJournal.get_sheet_by_name('IMGJournalData')

getFirstPageFromSheet(sheetMain, sheetJournal)
wbMain.save('SCArticles&JACCJournals1.xlsx')
print('Done with IMG')
time.sleep(10)

# HF
wbJournal = openpyxl.load_workbook('WebsiteToText\XHFJournalData.xlsx')
sheetJournal = wbJournal.get_sheet_by_name('HFJournalData')

getFirstPageFromSheet(sheetMain, sheetJournal)
wbMain.save('SCArticles&JACCJournals1.xlsx')
print('Done with HF')
time.sleep(10)

# EP
wbJournal = openpyxl.load_workbook('WebsiteToText\XEPJournalData.xlsx')
sheetJournal = wbJournal.get_sheet_by_name('EPJournalData')

getFirstPageFromSheet(sheetMain, sheetJournal)
wbMain.save('SCArticles&JACCJournals1.xlsx')
print('Done with EP')
time.sleep(10)

# BTS
wbJournal = openpyxl.load_workbook('WebsiteToText\XBTSJournalData.xlsx')
sheetJournal = wbJournal.get_sheet_by_name('BTSJournalData')

getFirstPageFromSheet(sheetMain, sheetJournal)
wbMain.save('SCArticles&JACCJournals1.xlsx')
print('Done with BTS')
time.sleep(25)
